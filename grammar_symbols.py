#!/usr/bin/env python3
import argparse

import openpyxl
import random
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

"""
Script to generated R-loop words as specified by a BED file.


Copyright 2020 Margherita Maria Ferrari.


This file is part of GrammarSymbols.

GrammarSymbols is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

GrammarSymbols is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with GrammarSymbols.  If not, see <http://www.gnu.org/licenses/>.
"""


class GrammarSymbols:
    __alpha = '\u03B1'
    __beta = '\u03B2'
    __delta = '\u03B4'
    __rho = '\u03C1'
    __rho_hat = __rho + '^'
    __sigma = '\u03C3'
    __sigma_hat = __sigma + '^'
    __tau = '\u03C4'
    __tau_hat = __tau + '^'
    __omega = '\u03C9'

    GREEK_MAPPING_R1 = {
        'W1': __sigma_hat,
        'W2': __sigma_hat,
        'W3': __sigma,
        'W4': __sigma,
        'N1': __sigma_hat,  # after R-loop
        'N2': __rho,
        'N3': __sigma  # before R-loop
    }

    GREEK_MAPPING_R2 = {
        'W1': __tau,
        'W2': __tau,
        'W3': __tau_hat,
        'W4': __tau_hat,
        'N1': __tau,  # after R-loop
        'N2': __rho,
        'N3': __tau_hat  # before R-loop
    }

    GREEK_MAPPING_R3 = {
        'W1': __sigma_hat,
        'W2': __sigma_hat,
        'W3': __sigma,
        'W4': __sigma,
        'N1': __sigma_hat,  # after R-loop
        'N2': __rho,
        'N3': __sigma  # before R-loop
    }

    @classmethod
    def __get_order_key(cls, x):
        if not x:
            return 0

        return sum([int(i) for i in x.split('_')])

    @classmethod
    def __get_regions(cls, gene_seq, bed_start, bed_end):
        if not gene_seq:
            raise AssertionError('Specify a gene sequence')
        if bed_start > bed_end:
            raise AssertionError('Start index must be lower or equal than end index')
        if len(gene_seq) < bed_end:
            raise AssertionError('End index too large')

        return gene_seq[:bed_start], gene_seq[bed_start:bed_end], gene_seq[bed_end:]

    @classmethod
    def __find_locations(cls, val, region1_values, region2_values, region3_values, region4_values, ret):
        if val in ret.keys():
            return ret.get(val, list())

        tmp = dict()  # key = weight, value = [W1_..., W2_...] --- a seq may have the same weight in more region
        if val in region1_values.keys():
            tmp_key = region1_values.get(val, 0)
            if tmp_key not in tmp.keys():
                tmp[tmp_key] = list()
            tmp[tmp_key].append('W1_' + str(tmp_key))
        if val in region2_values.keys():
            tmp_key = region2_values.get(val, 0)
            if tmp_key not in tmp.keys():
                tmp[tmp_key] = list()
            tmp[tmp_key].append('W2_' + str(tmp_key))
        if val in region3_values.keys():
            tmp_key = region3_values.get(val, 0)
            if tmp_key not in tmp.keys():
                tmp[tmp_key] = list()
            tmp[tmp_key].append('W3_' + str(tmp_key))
        if val in region4_values.keys():
            tmp_key = region4_values.get(val, 0)
            if tmp_key not in tmp.keys():
                tmp[tmp_key] = list()
            tmp[tmp_key].append('W4_' + str(tmp_key))

        sorted_keys = list(tmp.keys())
        sorted_keys.sort(reverse=True)
        sorted_values = list()

        for k in sorted_keys:
            sorted_values.extend(tmp[k])

        ret[val] = sorted_values
        return sorted_values

    @classmethod
    def __count_occurrences(cls, val, data, ret):
        if val in ret.keys():
            return ret.get(val, list())

        r1_count = 0
        r2_count = 0
        r3_count = 0

        for v in data.values():
            r1_count += v.get('r1', list()).count(val)
            r2_count += v.get('r2_rev', list()).count(val)
            r3_count += v.get('r3_rev', list()).count(val)

        tmp = dict()
        tmp[r1_count] = list()
        tmp[r1_count].append('N1_' + str(r1_count))

        if r2_count not in tmp.keys():
            tmp[r2_count] = list()
        tmp[r2_count].append('N2_' + str(r2_count))

        if r3_count not in tmp.keys():
            tmp[r3_count] = list()
        tmp[r3_count].append('N3_' + str(r3_count))

        sorted_keys = list(tmp.keys())
        sorted_keys.sort(reverse=True)
        sorted_values = list()

        for k in sorted_keys:
            sorted_values.extend(tmp[k])

        ret[val] = sorted_values
        return sorted_values

    @classmethod
    def get_args(cls):
        parser = argparse.ArgumentParser(description='Regions extractor')
        parser.add_argument('-f', '--input-fasta', metavar='FASTA_IN_FILE', type=str, required=True,
                            help='FASTA input file', default=None)
        parser.add_argument('-b', '--input-bed', metavar='BED_IN_FILE', type=str, required=True,
                            help='BED input file', default=None)
        parser.add_argument('-x', '--input-xlsx', metavar='XLSX_IN_FILE', type=str, required=False,
                            help='XLSX input file', default=None)
        parser.add_argument('-s', '--start-index', metavar='START_INDEX', type=int, required=True,
                            help='Start index of gene region', default=0)
        parser.add_argument('-e', '--end-index', metavar='END_INDEX', type=int, required=True,
                            help='End index of gene region', default=0)
        parser.add_argument('-n', '--num-rloops', metavar='NUM_RLOOPS', type=int, required=False,
                            help='Consider only NUM_RLOOPS rloops inside the BED file', default=10)
        parser.add_argument('-r', '--random-rloops', required=False, action='store_true',
                            help='Consider only NUM_RLOOPS random rloops inside the BED file')
        parser.add_argument('-w', '--window-length', metavar='WINDOW_LENGTH', type=int, required=False,
                            help='Number of nucleotides in single region', default=5)
        parser.add_argument('-o', '--output-file', metavar='OUTPUT_FILE', type=str, required=False,
                            help='Output XLSX file', default='output')
        return parser.parse_args()

    @classmethod
    def extract_regions(cls, fasta_in, bed_in, start_idx, end_idx, window_length=5, max_rloops=1800,
                        random_rloops=False, xlsx_in=None, out_file='output'):
        res = dict()
        with open(fasta_in, 'r') as fin:
            fin.readline()
            gene_seq = fin.readline().strip().upper()

        gene_seq = gene_seq[start_idx:end_idx]

        with open(bed_in, 'r') as fin:
            line = fin.readline()

            i = 1  # We keep track of the row we are reading in the BED file
            while line:
                parts = line.strip().split('\t')
                idx_1 = int(parts[1])
                idx_2 = int(parts[2])

                r1, r2, r3 = cls.__get_regions(gene_seq, idx_1 - start_idx, idx_2 - start_idx)
                r1_rev = r1[::-1]
                r2_rev = r2[::-1]
                r3_rev = r3[::-1]

                # res contains info from all the R-loops in BED file
                res[str(idx_1) + '_' + str(idx_2) + '_' + str(i)] = {'r1': [r1[i:i + window_length] for
                                                                            i in range(0, len(r1), window_length)],
                                                                     'r1_rev': [r1_rev[i:i + window_length][::-1] for i
                                                                                in range(0, len(r1), window_length)],
                                                                     'r2': [r2[i:i + window_length] for
                                                                            i in range(0, len(r2), window_length)],
                                                                     'r2_rev': [r2_rev[i:i + window_length][::-1] for i
                                                                                in range(0, len(r2), window_length)],
                                                                     'r3': [r3[i:i + window_length] for
                                                                            i in range(0, len(r3), window_length)],
                                                                     'r3_rev': [r3_rev[i:i + window_length][::-1] for i
                                                                                in range(0, len(r3), window_length)]
                                                                     }

                line = fin.readline()
                i += 1

        if random_rloops:
            sample_size = max_rloops if max_rloops <= len(res.keys()) else len(res.keys())
            res = {i: res[i] for i in random.sample(res.keys(), sample_size)}

        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Grammar Symbols')
        ws.append(('Gene', str(start_idx) + '-' + str(end_idx), gene_seq))
        ws.append(())
        ws.append(('Note: R-loop coordinates are given wrt full plasmid',))
        ws.append(())

        max_cols = max_rloops * 9  # max_cols: to print only n R-loops
        sorted_keys = list(res.keys())  # Need sorted keys bc we iterate on them (if not ordered, result not consistent)
        sorted_keys.sort(key=cls.__get_order_key)
        header = list()

        for k in sorted_keys:
            header.extend((k + '_r1', k + '_r1_extra', k + '_r1_funny_letters', k + '_r2', k + '_r2_extra',
                           k + '_r2_funny_letters', k + '_r3', k + '_r3_extra', k + '_r3_funny_letters'))

            if len(header) >= max_cols:  # Check we do not have more than max cols allowed by Excel
                break

        ws.append(header)

        wb_region1_values = dict()
        wb_region2_values = dict()
        wb_region3_values = dict()
        wb_region4_values = dict()
        wb_region2_extra_values = dict()
        wb_region3_extra_values = dict()

        if xlsx_in:
            wb_regions = openpyxl.load_workbook(xlsx_in, read_only=True)
            for ws_region in wb_regions.worksheets:
                if ws_region.title.endswith('1'):
                    wb_region1_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
                elif ws_region.title.endswith('2'):
                    wb_region2_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
                elif ws_region.title.endswith('3'):
                    wb_region3_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
                elif ws_region.title.endswith('4'):
                    wb_region4_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
                elif ws_region.title.endswith('2 Extra'):
                    wb_region2_extra_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
                elif ws_region.title.endswith('3 Extra'):
                    wb_region3_extra_values = {x[0].value: x[7].value for x in list(ws_region.rows)}
            wb_regions.close()

        i = 0
        word_dict = dict()
        last_cells = dict()
        row = list()
        locations = dict()
        counts = dict()

        while len(row) > 0 or i == 0:
            j = 3
            row = list()
            use_red_text = False

            # We read one value from r1, r2 and r3, at the same time but they have different lengths
            for k in sorted_keys:  # This "for" generates a row in the output file
                if not word_dict.get(k, None):
                    word_dict[k] = {'r1_funny_letters': list(), 'r2_funny_letters': list(), 'r3_funny_letters': list()}

                use_red_text = not use_red_text
                cell_font = Font(color='FF0000') if use_red_text else Font(color='000000')

                r1_val = res[k]['r1'][i] if len(res[k]['r1']) > i else None
                r1_val_locations = cls.__find_locations(r1_val, wb_region1_values, wb_region2_values, wb_region3_values,
                                                        wb_region4_values, locations) if r1_val else None
                # if r1_val not in "most relevant list"
                if (r1_val_locations is None or len(r1_val_locations) < 1) and r1_val:
                    r1_val_locations = cls.__count_occurrences(r1_val, res, counts)

                r2_val = res[k]['r2_rev'][i] if len(res[k]['r2_rev']) > i else None
                r2_val_locations = cls.__find_locations(r2_val, wb_region1_values, wb_region2_values, wb_region3_values,
                                                        wb_region4_values, locations) if r2_val else None

                if (r2_val_locations is None or len(r2_val_locations) < 1) and r2_val:
                    r2_val_locations = cls.__count_occurrences(r2_val, res, counts)

                r3_val = res[k]['r3_rev'][i] if len(res[k]['r3_rev']) > i else None
                r3_val_locations = cls.__find_locations(r3_val, wb_region1_values, wb_region2_values, wb_region3_values,
                                                        wb_region4_values, locations) if r3_val else None

                if (r3_val_locations is None or len(r3_val_locations) < 1) and r3_val:
                    r3_val_locations = cls.__count_occurrences(r3_val, res, counts)

                r1_funny_letters = None
                if r1_val_locations and len(r1_val_locations) > 0:
                    tmp_val = r1_val_locations[0].split('_')[1]  # take weight/count
                    # collect N/W whose value is the same as the maximum one
                    tmp_keys = [i.split('_')[0].upper() for i in r1_val_locations if i.split('_')[1] == tmp_val]

                    if 'N2' in tmp_keys:
                        tmp_locs = cls.__find_locations(r1_val, dict(), wb_region2_extra_values,
                                                        wb_region3_extra_values, dict(), dict())
                        if len(tmp_locs) > 0:
                            while 'N2' in tmp_keys:
                                tmp_keys.remove('N2')

                            tmp_val = tmp_locs[0].split('_')[1]
                            tmp_keys.extend([i.split('_')[0].upper() for i in tmp_locs if i.split('_')[1] == tmp_val])

                    r1_funny_letters = list(set([cls.GREEK_MAPPING_R1.get(i, '?') for i in set(tmp_keys)]))

                    if cls.__tau in r1_funny_letters and cls.__tau_hat in r1_funny_letters:
                        r1_funny_letters.remove(cls.__tau)
                        r1_funny_letters.remove(cls.__tau_hat)
                        r1_funny_letters.append(cls.__delta)

                    if cls.__sigma in r1_funny_letters and cls.__sigma_hat in r1_funny_letters:
                        r1_funny_letters.remove(cls.__sigma)
                        r1_funny_letters.remove(cls.__sigma_hat)
                        r1_funny_letters.append(cls.__beta)

                    if len(r1_funny_letters) == 1:
                        word_dict[k]['r1_funny_letters'].append(r1_funny_letters[0])
                    else:
                        word_dict[k]['r1_funny_letters'].append('(' + ','.join(r1_funny_letters) + ')')

                    x = last_cells.get(k, dict())
                    x['r1'] = str(j) + '_' + str(i + 6)
                    last_cells[k] = x

                r2_funny_letters = None
                if r2_val_locations and len(r2_val_locations) > 0:
                    tmp_val = r2_val_locations[0].split('_')[1]
                    tmp_keys = [i.split('_')[0].upper() for i in r2_val_locations if i.split('_')[1] == tmp_val]

                    if 'N2' in tmp_keys:
                        tmp_locs = cls.__find_locations(r2_val, dict(), wb_region2_extra_values,
                                                        wb_region3_extra_values, dict(), dict())
                        if len(tmp_locs) > 0:
                            while 'N2' in tmp_keys:
                                tmp_keys.remove('N2')

                            tmp_val = tmp_locs[0].split('_')[1]
                            tmp_keys.extend([i.split('_')[0].upper() for i in tmp_locs if i.split('_')[1] == tmp_val])

                    r2_funny_letters = list(set([cls.GREEK_MAPPING_R2.get(i, '?') for i in set(tmp_keys)]))

                    if cls.__tau in r2_funny_letters and cls.__tau_hat in r2_funny_letters:
                        r2_funny_letters.remove(cls.__tau)
                        r2_funny_letters.remove(cls.__tau_hat)
                        r2_funny_letters.append(cls.__delta)

                    if cls.__sigma in r2_funny_letters and cls.__sigma_hat in r2_funny_letters:
                        r2_funny_letters.remove(cls.__sigma)
                        r2_funny_letters.remove(cls.__sigma_hat)
                        r2_funny_letters.append(cls.__beta)

                    if len(r2_funny_letters) == 1:
                        word_dict[k]['r2_funny_letters'].append(r2_funny_letters[0])
                    else:
                        word_dict[k]['r2_funny_letters'].append('(' + ','.join(r2_funny_letters) + ')')

                r3_funny_letters = None
                if r3_val_locations and len(r3_val_locations) > 0:
                    tmp_val = r3_val_locations[0].split('_')[1]
                    tmp_keys = [i.split('_')[0].upper() for i in r3_val_locations if i.split('_')[1] == tmp_val]

                    if 'N2' in tmp_keys:
                        tmp_locs = cls.__find_locations(r3_val, dict(), wb_region2_extra_values,
                                                        wb_region3_extra_values, dict(), dict())
                        if len(tmp_locs) > 0:
                            while 'N2' in tmp_keys:
                                tmp_keys.remove('N2')

                            tmp_val = tmp_locs[0].split('_')[1]
                            tmp_keys.extend([i.split('_')[0].upper() for i in tmp_locs if i.split('_')[1] == tmp_val])

                    r3_funny_letters = list(set([cls.GREEK_MAPPING_R3.get(i, '?') for i in set(tmp_keys)]))

                    if cls.__tau in r3_funny_letters and cls.__tau_hat in r3_funny_letters:
                        r3_funny_letters.remove(cls.__tau)
                        r3_funny_letters.remove(cls.__tau_hat)
                        r3_funny_letters.append(cls.__delta)

                    if cls.__sigma in r3_funny_letters and cls.__sigma_hat in r3_funny_letters:
                        r3_funny_letters.remove(cls.__sigma)
                        r3_funny_letters.remove(cls.__sigma_hat)
                        r3_funny_letters.append(cls.__beta)

                    if len(r3_funny_letters) == 1:
                        word_dict[k]['r3_funny_letters'].append(r3_funny_letters[0])
                    else:
                        word_dict[k]['r3_funny_letters'].append('(' + ','.join(r3_funny_letters) + ')')

                    x = last_cells.get(k, dict())
                    x['r3'] = str(j + 6) + '_' + str(i + 6)
                    last_cells[k] = x

                if r1_val:
                    r1_val = WriteOnlyCell(ws, value=r1_val)
                    r1_val.font = cell_font

                if r1_val_locations:
                    r1_val_location = WriteOnlyCell(ws, value=', '.join(r1_val_locations))
                    r1_val_location.font = cell_font
                else:
                    r1_val_location = None

                if r1_funny_letters:
                    r1_funny_letter = WriteOnlyCell(ws, value=', '.join(r1_funny_letters))
                    r1_funny_letter.font = cell_font
                else:
                    r1_funny_letter = None

                if r2_val:
                    r2_val = WriteOnlyCell(ws, value=r2_val)
                    r2_val.font = cell_font

                if r2_val_locations:
                    r2_val_location = WriteOnlyCell(ws, value=', '.join(r2_val_locations))
                    r2_val_location.font = cell_font
                else:
                    r2_val_location = None

                if r2_funny_letters:
                    r2_funny_letter = WriteOnlyCell(ws, value=', '.join(r2_funny_letters))
                    r2_funny_letter.font = cell_font
                else:
                    r2_funny_letter = None

                if r3_val:
                    r3_val = WriteOnlyCell(ws, value=r3_val)
                    r3_val.font = cell_font

                if r3_val_locations:
                    r3_val_location = WriteOnlyCell(ws, value=', '.join(r3_val_locations))
                    r3_val_location.font = cell_font
                else:
                    r3_val_location = None

                if r3_funny_letters:
                    r3_funny_letter = WriteOnlyCell(ws, value=', '.join(r3_funny_letters))
                    r3_funny_letter.font = cell_font
                else:
                    r3_funny_letter = None

                row.extend((r1_val, r1_val_location, r1_funny_letter,
                            r2_val, r2_val_location, r2_funny_letter,
                            r3_val, r3_val_location, r3_funny_letter))

                j += 9

                if len(row) >= max_cols:  # Check we do not have more than max cols allowed by Excel
                    break

            # If we have reached the end of r1, r2, r3 in res, we have None everywhere in row and we have to clear it
            if len([i for i in row if i is not None]) < 1:
                row.clear()

            ws.append(row)
            i += 1

        ws.close()
        wb.save(out_file)

        wb = openpyxl.load_workbook(out_file)
        ws = wb['Grammar Symbols']

        for k, x, y in [(i, j.get('r1', None), j.get('r3', None)) for i, j in last_cells.items()]:
            x_col = int(x.split('_')[0])
            x_row = int(x.split('_')[1])
            v = ws.cell(row=x_row, column=x_col - 2).value
            cell_value = cls.__omega + str(len(v))

            if len(v) == window_length:
                cell_value = ws.cell(row=x_row, column=x_col).value + cls.__omega + '0'

            ws.cell(row=x_row, column=x_col).value = cell_value
            last_cells[k]['r1'] = cell_value

            y_col = int(y.split('_')[0])
            y_row = int(y.split('_')[1])
            v = ws.cell(row=y_row, column=y_col - 2).value
            cell_value = cls.__alpha + str(len(v))

            if len(v) == window_length:
                cell_value = cls.__alpha + '0' + ws.cell(row=y_row, column=y_col).value

            ws.cell(row=y_row, column=y_col).value = cell_value
            last_cells[k]['r3'] = cell_value

        wb.save(out_file)

        with open(out_file + '.txt', 'w') as fout:
            for k, v in word_dict.items():
                v['r1_funny_letters'][-1] = last_cells.get(k, dict()).get('r1', cls.__omega)
                v['r3_funny_letters'][-1] = last_cells.get(k, dict()).get('r3', cls.__alpha)
                line = k + ': ' + ''.join(v['r1_funny_letters']) + ''.join(reversed(v['r2_funny_letters'])) + \
                       ''.join(reversed(v['r3_funny_letters']))
                fout.write(line + '\n')


if __name__ == '__main__':
    args = vars(GrammarSymbols.get_args())
    GrammarSymbols.extract_regions(args.get('input_fasta', None), args.get('input_bed', None),
                                   args.get('start_index', 0), args.get('end_index', 0), args.get('window_length', 5),
                                   args.get('num_rloops', 10), args.get('random_rloops', False),
                                   args.get('input_xlsx', None), args.get('output_file', 'output'))
